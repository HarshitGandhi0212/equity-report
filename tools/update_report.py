#!/usr/bin/env python3
"""
update_report.py · v2
─────────────────────────────────────────────────────────────────────────
Reads a Screener.in Excel export → produces the JSON data block for a
stock's risk report.

Tested against Screener template version 2.1 (May 2026).

USAGE:
    python tools/update_report.py biocon              # preview JSON (dry run)
    python tools/update_report.py biocon --in-place   # patch reports/biocon.html
    python tools/update_report.py biocon --excel data/Biocon.xlsx
    python tools/update_report.py biocon --merge      # merge with existing JSON
                                                       (preserves judgment fields)

KEY DESIGN DECISIONS:
  1. Uses 'Adjusted Equity Shares in Cr' (Screener's DERIVED row) for all
     per-share calculations. This is bonus-adjusted automatically.
  2. EBITDA = PBT + Interest + Depreciation - Other Income
     (i.e. operating EBITDA, excluding Other Income — pure ops)
  3. Never overwrites judgment fields (scores, verdict, catalysts, etc.)
     in --merge mode. Only adds/refreshes the numeric data.
  4. Surfaces a list of "fields still pending" so you don't accidentally
     ship a report with stale shareholding data or missing 52W range.

WHAT THIS SCRIPT CANNOT DO:
  - Read shareholding pattern (Screener Excel doesn't export it)
  - Get intraday 52W high/low (only annual closes are in the Excel)
  - Get beta, ROCE, segment revenue (not in default Screener template)
  - Write the analysis (verdict, catalysts, risks — those are your job)

Install:
    pip install openpyxl
"""

import argparse
import json
import re
import sys
from datetime import date, timedelta, datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ─────────────────────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
DATA_DIR = REPO_ROOT / "data"
REPORTS_DIR = REPO_ROOT / "reports"


# ─────────────────────────────────────────────────────────────
# PER-STOCK STATIC CONFIG (extend as you add stocks)
# ─────────────────────────────────────────────────────────────

STOCK_CONFIG = {
    "biocon": {
        "ticker": "BIOCON",
        "name": "Biocon Ltd",
        "exchange": "NSE",
        "code": "532523",
        "bseCode": "532523",
        "isin": "INE376G01013",
        "sector": "Pharma · Biotech",
        "subSector": "Innovation-led Global Biopharma",
        "index": "Nifty Midcap 150 · Nifty Pharma",
        "screenerSlug": "BIOCON",
        "bseSlug": "biocon-ltd/biocon",
        "investorRelations": "https://www.biocon.com/investor-relations/",
    },
    "hfcl": {
        "ticker": "HFCL",
        "name": "HFCL Ltd",
        "exchange": "NSE",
        "code": "500183",
        "bseCode": "500183",
        "isin": "INE548A01028",
        "sector": "Telecom Equipment",
        "subSector": "Telecom · OFC · Defence Electronics",
        "index": "Nifty Smallcap 100",
        "screenerSlug": "HFCL",
        "bseSlug": "hfcl-ltd/hfcl",
        "investorRelations": "https://www.hfcl.com/investor-relations/",
    },
    "jainrec": {
        "ticker": "JAINREC",
        "name": "Jain Resource Recycling Ltd",
        "exchange": "NSE",
        "code": "544537",
        "bseCode": "544537",
        "isin": "INE0YD401026",
        "sector": "Metals & Mining",
        "subSector": "Non-Ferrous Metal Recycling",
        "index": "Smallcap · Listed Oct 2025",
        "screenerSlug": "JAINREC",
        "bseSlug": "jain-resource-recycling-ltd/jainrec",
        "investorRelations": "",
    },
    "tiindia": {
        "ticker": "TIINDIA",
        "name": "Tube Investments of India Ltd",
        "exchange": "NSE",
        "code": "540762",
        "bseCode": "540762",
        "isin": "INE974X01010",
        "sector": "Auto Ancillary",
        "subSector": "Auto Anc · Engineering · EV",
        "index": "Nifty Midcap · Murugappa Group",
        "screenerSlug": "TIINDIA",
        "bseSlug": "tube-investments-of-india-ltd/tiindia",
        "investorRelations": "https://www.tiindia.com/investor-relations",
    },
}


# ─────────────────────────────────────────────────────────────
# SCREENER EXCEL READER
# ─────────────────────────────────────────────────────────────

class ScreenerExcel:
    """Reads a Screener.in Data Sheet, exposes section-aware row lookups."""

    SECTION_HEADERS = {
        "META", "PROFIT & LOSS", "Quarters", "BALANCE SHEET",
        "CASH FLOW:", "PRICE:", "DERIVED:"
    }

    def __init__(self, path: Path):
        self.path = path
        self.wb = load_workbook(path, data_only=True, read_only=True)
        if "Data Sheet" not in self.wb.sheetnames:
            raise ValueError(
                f"'Data Sheet' tab not found in {path.name}. "
                f"Sheets present: {self.wb.sheetnames}"
            )
        self.ds = self.wb["Data Sheet"]
        self._parse()

    def _parse(self):
        """Build a section → {label: [values]} map."""
        self.sections = {"TOP": {}}
        current = "TOP"
        for row in range(1, self.ds.max_row + 1):
            label = self.ds.cell(row=row, column=1).value
            if not label:
                continue
            label = str(label).strip()
            vals = [self.ds.cell(row=row, column=c).value
                    for c in range(2, self.ds.max_column + 1)]

            if label in self.SECTION_HEADERS:
                current = label
                self.sections[current] = {}
                if any(v is not None and v != "" for v in vals):
                    self.sections[current][label] = vals
            else:
                self.sections.setdefault(current, {})[label] = vals

    def get(self, section, label, default=None):
        return self.sections.get(section, {}).get(label, default or [])

    def clean(self, section, label):
        vals = self.get(section, label, [])
        return [v for v in vals if v is not None and v != ""]

    def latest(self, section, label):
        vals = self.clean(section, label)
        return vals[-1] if vals else None

    def prior(self, section, label, n=1):
        vals = self.clean(section, label)
        if len(vals) > n:
            return vals[-(n + 1)]
        return None

    @property
    def company_name(self):
        return self.sections.get("TOP", {}).get("COMPANY NAME", [None])[0]

    @staticmethod
    def excel_to_date(v):
        if isinstance(v, datetime):
            return v.date().isoformat()
        if isinstance(v, (int, float)):
            return (date(1899, 12, 30) + timedelta(days=int(v))).isoformat()
        return None


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────

def safe_div(a, b):
    if a is None or b is None or b == 0:
        return None
    try:
        return a / b
    except (TypeError, ZeroDivisionError):
        return None


def safe_pct(a, b, decimals=2):
    v = safe_div(a, b)
    if v is None:
        return None
    return round(v * 100, decimals)


def compute_ebitda(pbt, interest, depreciation, other_income):
    """EBITDA = PBT + Interest + Depreciation - Other Income (operating EBITDA)."""
    if any(v is None for v in [pbt, interest, depreciation, other_income]):
        return None
    return round(pbt + interest + depreciation - other_income, 2)


# ─────────────────────────────────────────────────────────────
# BUILD DATA BLOCK
# ─────────────────────────────────────────────────────────────

def build_data_block(s: ScreenerExcel, stock_key: str) -> dict:
    cfg = STOCK_CONFIG[stock_key]
    today = date.today()

    # Raw values
    cmp_val = s.latest("META", "Current Price")
    mcap = s.latest("META", "Market Capitalization")

    sales = s.latest("PROFIT & LOSS", "Sales")
    sales_prior = s.prior("PROFIT & LOSS", "Sales", 1)
    pat = s.latest("PROFIT & LOSS", "Net profit")
    pat_prior = s.prior("PROFIT & LOSS", "Net profit", 1)
    dividend_amt = s.latest("PROFIT & LOSS", "Dividend Amount")

    pbt = s.latest("PROFIT & LOSS", "Profit before tax")
    interest = s.latest("PROFIT & LOSS", "Interest")
    depr = s.latest("PROFIT & LOSS", "Depreciation")
    other_inc = s.latest("PROFIT & LOSS", "Other Income")

    pbt_p = s.prior("PROFIT & LOSS", "Profit before tax", 1)
    interest_p = s.prior("PROFIT & LOSS", "Interest", 1)
    depr_p = s.prior("PROFIT & LOSS", "Depreciation", 1)
    other_inc_p = s.prior("PROFIT & LOSS", "Other Income", 1)

    equity_capital = s.latest("BALANCE SHEET", "Equity Share Capital")
    reserves = s.latest("BALANCE SHEET", "Reserves")
    borrowings = s.latest("BALANCE SHEET", "Borrowings")

    adj_shares = s.latest("DERIVED:", "Adjusted Equity Shares in Cr")
    adj_shares_prior = s.prior("DERIVED:", "Adjusted Equity Shares in Cr", 1)

    cfo = s.latest("CASH FLOW:", "Cash from Operating Activity")

    pl_dates = s.clean("PROFIT & LOSS", "Report Date")
    latest_period = s.excel_to_date(pl_dates[-1]) if pl_dates else None
    prior_period = s.excel_to_date(pl_dates[-2]) if len(pl_dates) >= 2 else None

    # Derived
    net_worth = (equity_capital or 0) + (reserves or 0)
    eps_latest = safe_div(pat, adj_shares)
    eps_prior = safe_div(pat_prior, adj_shares_prior)
    book_value = safe_div(net_worth, adj_shares)
    pe = safe_div(cmp_val, eps_latest)
    pb = safe_div(cmp_val, book_value)
    debt_equity = safe_div(borrowings, net_worth)
    roe_pct = safe_pct(pat, net_worth)
    revenue_yoy = safe_pct(sales - sales_prior, sales_prior) if sales and sales_prior else None
    pat_yoy = safe_pct(pat - pat_prior, pat_prior) if pat and pat_prior else None
    mcap_sales = safe_div(mcap, sales)

    ebitda = compute_ebitda(pbt, interest, depr, other_inc)
    ebitda_prior = compute_ebitda(pbt_p, interest_p, depr_p, other_inc_p)
    ebitda_margin = safe_pct(ebitda, sales)
    ebitda_yoy = safe_pct(ebitda - ebitda_prior, ebitda_prior) if ebitda and ebitda_prior else None
    cfo_ebitda = safe_pct(cfo, ebitda)

    div_per_share = safe_div(dividend_amt, adj_shares)
    div_yield = safe_pct(div_per_share, cmp_val) if div_per_share else None

    block = {
        "stock": {
            "ticker": cfg["ticker"],
            "name": cfg["name"],
            "exchange": cfg["exchange"],
            "code": cfg["code"],
            "bseCode": cfg["bseCode"],
            "isin": cfg["isin"],
            "sector": cfg["sector"],
            "subSector": cfg["subSector"],
            "index": cfg["index"],
        },
        "asOf": {
            "dataDate": today.isoformat(),
            "dataDateDisplay": today.strftime("%d-%b-%Y"),
            "latestReportPeriod": latest_period,
            "priorReportPeriod": prior_period,
            "analysisDate": "_PENDING_MANUAL_",
            "analysisDateDisplay": "_PENDING_MANUAL_",
            "source": "Screener.in consolidated (Excel export)",
            "screenerExportFile": s.path.name,
            "nextScheduledUpdate": "_PENDING_MANUAL_",
            "version": "auto-generated",
        },
        "price": {
            "cmp": cmp_val,
            "change": None,
            "changePct": None,
            "high52": None,
            "low52": None,
            "yearReturn1y": None,
            "beta": None,
        },
        "valuation": {
            "marketCapCr": round(mcap, 2) if mcap else None,
            "pe": round(pe, 2) if pe else None,
            "pb": round(pb, 2) if pb else None,
            "evEbitda": None,
            "peg": None,
            "dividendYield": round(div_yield, 2) if div_yield else None,
            "mcapSales": round(mcap_sales, 2) if mcap_sales else None,
            "epsTtm": round(eps_latest, 2) if eps_latest else None,
            "epsPriorYear": round(eps_prior, 2) if eps_prior else None,
            "bookValue": round(book_value, 2) if book_value else None,
        },
        "fundamentals": {
            "revenueTtmCr": round(sales, 2) if sales else None,
            "revenueYoyPct": revenue_yoy,
            "ebitdaTtmCr": ebitda,
            "ebitdaYoyPct": ebitda_yoy,
            "ebitdaMarginPct": ebitda_margin,
            "patTtmCr": round(pat, 2) if pat else None,
            "patYoyPct": pat_yoy,
            "patPriorCr": round(pat_prior, 2) if pat_prior else None,
        },
        "health": {
            "debtEquity": round(debt_equity, 2) if debt_equity else None,
            "interestCoverage": None,
            "cfoEbitdaPct": cfo_ebitda,
            "promoterHolding": None,
            "fiiHolding": None,
            "diiHolding": None,
            "publicHolding": None,
            "roe": roe_pct,
            "roce": None,
            "promoterPledgePct": None,
        },
        "scores": {
            "_note": "JUDGMENT FIELDS — not auto-updated",
            "valuation": None,
            "valuationFlag": None,
            "financialHealth": None,
            "financialHealthFlag": None,
            "growth": None,
            "growthFlag": None,
            "composite": None,
            "compositeBand": None,
            "compositeBandColor": None,
        },
        "verdict": {
            "_note": "JUDGMENT FIELDS — not auto-updated",
            "rating": None,
            "ratingPosition": None,
            "fairValue": None,
            "entryZone": None,
            "stopLoss": None,
            "positionCapPct": None,
            "consensusTp": None,
            "consensusUpsidePct": None,
        },
        "links": {
            "screener": f"https://www.screener.in/company/{cfg['screenerSlug']}/consolidated/",
            "nse": f"https://www.nseindia.com/get-quotes/equity?symbol={cfg['ticker']}",
            "bse": f"https://www.bseindia.com/stock-share-price/{cfg['bseSlug']}/{cfg['code']}/",
            "investorRelations": cfg.get("investorRelations", ""),
            "trendlyne": f"https://trendlyne.com/equity/Forecaster/{cfg['code']}/{cfg['ticker']}/",
        },
        "quarterly": _build_quarterly(s),
    }
    return block


def _build_quarterly(s: ScreenerExcel) -> list:
    q_dates = s.clean("Quarters", "Report Date")
    if not q_dates:
        return []

    sales_q = s.clean("Quarters", "Sales")
    op_profit_q = s.clean("Quarters", "Operating Profit")
    pat_q = s.clean("Quarters", "Net profit")
    pbt_q = s.clean("Quarters", "Profit before tax")

    quarters = []
    n = len(q_dates)
    start = max(0, n - 6)
    for i in range(start, n):
        sales = sales_q[i] if i < len(sales_q) else None
        op = op_profit_q[i] if i < len(op_profit_q) else None
        pat = pat_q[i] if i < len(pat_q) else None

        quarters.append({
            "period": ScreenerExcel.excel_to_date(q_dates[i]),
            "revenueCr": round(sales, 2) if sales else None,
            "operatingProfitCr": round(op, 2) if op else None,
            "operatingMarginPct": safe_pct(op, sales) if op and sales else None,
            "patCr": round(pat, 2) if pat else None,
        })
    return quarters


# ─────────────────────────────────────────────────────────────
# MERGE WITH EXISTING (preserves judgment fields)
# ─────────────────────────────────────────────────────────────

def merge_with_existing(new_data: dict, existing_path: Path) -> dict:
    if not existing_path.exists():
        return new_data

    content = existing_path.read_text(encoding="utf-8")
    m = re.search(r'<script id="reportData"[^>]*>(.*?)</script>', content, re.DOTALL)
    if not m:
        return new_data

    try:
        existing = json.loads(m.group(1))
    except json.JSONDecodeError:
        print("Warning: existing JSON is malformed; ignoring", file=sys.stderr)
        return new_data

    # Preserve judgment
    if "scores" in existing:
        new_data["scores"] = existing["scores"]
    if "verdict" in existing:
        new_data["verdict"] = existing["verdict"]

    # Preserve shareholding + ratios not in Excel
    for f in ["promoterHolding", "fiiHolding", "diiHolding",
              "publicHolding", "roce", "interestCoverage",
              "promoterPledgePct"]:
        if new_data["health"].get(f) is None:
            new_data["health"][f] = existing.get("health", {}).get(f)

    # Preserve 52W + beta
    for f in ["high52", "low52", "yearReturn1y", "beta"]:
        if new_data["price"].get(f) is None:
            new_data["price"][f] = existing.get("price", {}).get(f)

    # Preserve EV/EBITDA, PEG
    for f in ["evEbitda", "peg"]:
        if new_data["valuation"].get(f) is None:
            new_data["valuation"][f] = existing.get("valuation", {}).get(f)

    # Preserve analysis dates
    for f in ["analysisDate", "analysisDateDisplay", "nextScheduledUpdate"]:
        ev = existing.get("asOf", {}).get(f)
        if ev and ev != "_PENDING_MANUAL_":
            new_data["asOf"][f] = ev

    return new_data


# ─────────────────────────────────────────────────────────────
# HTML PATCH
# ─────────────────────────────────────────────────────────────

def patch_html(html_path: Path, data: dict) -> bool:
    if not html_path.exists():
        print(f"ERROR: HTML not found at {html_path}", file=sys.stderr)
        return False

    content = html_path.read_text(encoding="utf-8")
    pattern = re.compile(
        r'(<script id="reportData" type="application/json">)(.*?)(</script>)',
        re.DOTALL,
    )
    if not pattern.search(content):
        print(f"ERROR: <script id='reportData'> block not in {html_path}", file=sys.stderr)
        return False

    new_json = "\n" + json.dumps(data, indent=2, ensure_ascii=False) + "\n"
    new_content = pattern.sub(lambda m: m.group(1) + new_json + m.group(3), content)
    html_path.write_text(new_content, encoding="utf-8")
    return True


# ─────────────────────────────────────────────────────────────
# REPORTING
# ─────────────────────────────────────────────────────────────

def report_pending(data: dict):
    pending = []

    def walk(prefix, obj):
        if isinstance(obj, dict):
            for k, v in obj.items():
                if k.startswith("_"):
                    continue
                walk(f"{prefix}.{k}" if prefix else k, v)
        elif obj is None or obj == "_PENDING_MANUAL_":
            pending.append(prefix)

    walk("", data)

    if pending:
        print("\n⚠ Fields requiring manual fill:", file=sys.stderr)
        for p in pending:
            print(f"   • {p}", file=sys.stderr)
    else:
        print("\n✓ All fields populated.", file=sys.stderr)


def find_latest_excel(stock_key: str) -> Path:
    if not DATA_DIR.exists():
        raise FileNotFoundError(
            f"Data directory not found: {DATA_DIR}\nCreate it: mkdir -p {DATA_DIR}"
        )

    patterns = [f"{stock_key}*.xlsx", f"{stock_key.upper()}*.xlsx",
                f"{stock_key.capitalize()}*.xlsx"]
    candidates = []
    for pat in patterns:
        candidates.extend(DATA_DIR.glob(pat))
    candidates = sorted(set(candidates), key=lambda p: p.stat().st_mtime, reverse=True)

    if not candidates:
        raise FileNotFoundError(
            f"No Screener Excel found for '{stock_key}' in {DATA_DIR}/.\n"
            f"Download from screener.in and save as '{stock_key.capitalize()}.xlsx'."
        )
    return candidates[0]


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Update report data from Screener Excel export.",
    )
    parser.add_argument("stock", choices=list(STOCK_CONFIG.keys()))
    parser.add_argument("--excel", type=Path, default=None)
    parser.add_argument("--html", type=Path, default=None)
    parser.add_argument("--in-place", action="store_true")
    parser.add_argument("--merge", action="store_true")
    args = parser.parse_args()

    excel_path = args.excel or find_latest_excel(args.stock)
    print(f"Reading: {excel_path}", file=sys.stderr)

    screener = ScreenerExcel(excel_path)
    print(f"Company: {screener.company_name}", file=sys.stderr)

    data = build_data_block(screener, args.stock)

    html_path = args.html or (REPORTS_DIR / f"{args.stock}.html")
    if args.merge or args.in_place:
        data = merge_with_existing(data, html_path)

    if args.in_place:
        success = patch_html(html_path, data)
        if success:
            print(f"✓ Patched {html_path}", file=sys.stderr)
        report_pending(data)
        sys.exit(0 if success else 1)
    else:
        print(json.dumps(data, indent=2, ensure_ascii=False))
        report_pending(data)


if __name__ == "__main__":
    main()
